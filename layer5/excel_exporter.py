import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Exports analytics data into a multi-sheet Excel workbook using openpyxl.
    """

    def __init__(self, config: dict):
        cfg = config.get("reports", {})
        self.output_dir = cfg.get("output_dir", "data/reports")
        os.makedirs(self.output_dir, exist_ok=True)

    def generate(self, analytics, hours: int = 24, camera_id: str = "cam_0") -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return ""

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{camera_id}_{ts}.xlsx"
        path = os.path.join(self.output_dir, filename)

        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        def format_sheet(ws):
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 18
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # ── Sheet 1: Detections by Class ────────────────────────────────
        ws1 = wb.active
        ws1.title = "Detections by Class"
        ws1.append(["Class Name", "Detection Count"])
        
        data_class = analytics.detections_by_class(hours)
        for row in data_class:
            ws1.append([row["class_name"].title(), row["count"]])
        format_sheet(ws1)

        # ── Sheet 2: Alerts Summary ─────────────────────────────────────
        ws2 = wb.create_sheet(title="Alerts Summary")
        ws2.append(["Alert Type", "Count"])
        
        data_alerts = analytics.alerts_by_type(hours)
        for row in data_alerts:
            ws2.append([row["alert_type"].replace("_", " ").title(), row["count"]])
        format_sheet(ws2)

        # ── Sheet 3: Zone Dwell Times ───────────────────────────────────
        ws3 = wb.create_sheet(title="Zone Dwell Times")
        ws3.append(["Zone Name", "Average Dwell (s)", "Max Dwell (s)", "Total Visits"])
        
        data_dwell = analytics.avg_dwell_per_zone(hours//24 if hours >= 24 else 1)
        for row in data_dwell:
            ws3.append([
                row["zone_name"], 
                round(row["avg_seconds"], 2), 
                round(row["max_seconds"], 2), 
                row["visits"]
            ])
        format_sheet(ws3)

        # ── Sheet 4: Hourly Activity ────────────────────────────────────
        ws4 = wb.create_sheet(title="Hourly Activity")
        ws4.append(["Hour", "Detection Count"])
        
        data_hourly = analytics.detections_per_hour(hours//24 if hours >= 24 else 1)
        for row in data_hourly:
            ws4.append([row["hour"], row["count"]])
        format_sheet(ws4)

        wb.save(path)
        logger.info(f"Excel Export generated: {path}")
        return path
